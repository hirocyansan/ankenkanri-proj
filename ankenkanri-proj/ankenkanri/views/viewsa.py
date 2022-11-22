from asyncio.windows_events import NULL
# from pickle import NONE
# from re import I
# from unittest.mock import NonCallableMock
from django.shortcuts import render
from django.http import HttpResponse
from ..import forms
from django.forms import formset_factory
from django.utils import timezone
from django.contrib.auth.models import User
import datetime
from django.core.files.storage import FileSystemStorage

import sqlite3
import openpyxl
import os
# from numpy import true_divide
# from django import setup
from ..models import AnkenList
from ..models import AnkenStatus
from ..models import AnkenTorihikisaki
from ..models import AnkenTantosha
from ..models import AnkenKanjokamoku
from ..models import AnkenShiharaiPattern


# Create your views here.

# def index(request):
#     return HttpResponse('<h1> Good Bye </h1>')

def maxEdaban(kanriNo):
    edabanShitei=0
    anken = AnkenList.objects.order_by('hyojijun')
    for ankens in anken:
        
        if ankens.kanriNo == kanriNo and int(ankens.edaban) >= int(edabanShitei):
            edabanShitei = str(int(ankens.edaban)+1)
            
    return edabanShitei    


# def hyojijunRefresh():
#     anken = AnkenList.objects.order_by('hyojijun')
#     i = 1
#     for ankens in anken:
        
#         ankens.hyojijun = i
#         i=i+1    
            
    
    
    
    
    
    
    


def index (request):
    kazu = AnkenList.objects.all().count()
    context = {
        'ankenList' : AnkenList.objects.order_by('hyojijun'),
        'kazu' : kazu
    }
    request.session['sessionDisplayCode'] = 'dp00' 
    return render(request,'ankenkanri2/index.html',context)

def home(request):
    my_name = 'Taro Yamada'
    favourite_fruits = ['Apple','Grape','Lemon']
    my_info = {
        'name': 'Taro',
        'age': 18
    }
    return render(request, 'ankenkanri2/home.html',context={
        'my_name':my_name,
        'favourite_fruits':favourite_fruits,
        'my_info':my_info,
    })

def sample (request):
    name = 'ichiro yamada'
    height = 175.5
    weight = 72
    bmi = weight / (height /100)**2
    page_url = 'https://www.google.com'
    favourete_fruits = [
        'Apple','Grape','Lemon'
    ]
    msg = """hello
    my name is
    ichiro
    """
    msg2 = '1234567890'
    return render(request, 'ankenkanri2/sample.html',context = {
        'name':name,
        'bmi' : bmi,
        'page_url':page_url,
        'fruits':favourete_fruits,
        'msg':msg,
        'msg2':msg2
    })

class Country:
    
    def __init__(self, name, population, capital):
        self.name = name
        self.population = population
        self.capital = capital

def sample3(request):
    country = Country('Japan', 100000000, 'Tokyo')
    return render(request,'ankenkanri2/sample3.html',context={
        'country':country
    })
    
def export2 (request):
    kazu = AnkenList.objects.all().count()
    inputFile=request.POST.get('inputFile')
    fileShurui=request.POST.get('tableName')
    
    # print("ファイルの種類は",request.POST.get('tableName'))
    context = {
        'ankenList' : AnkenList.objects.order_by('hyojijun'),
        'inputFile' : inputFile,
        'kazu' : kazu,
        'shurui':fileShurui
    }
    if request.method == "POST":
        if "inputFile" in request.POST:
            
            if fileShurui=="案件管理データ":
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'].value = 'id'
                sheet['B1'].value = '経理承認'
                sheet['C1'].value = 'ステータスコード'
                sheet['D1'].value = '案件ステータス'
                sheet['E1'].value = '枝番グループ'
                sheet['F1'].value = '支払パターン'
                sheet['G1'].value = '管理No.'
                sheet['H1'].value = '枝番'
                sheet['I1'].value = '案件名'
                sheet['J1'].value = '取引先コード'
                sheet['K1'].value = '取引先名'
                sheet['L1'].value = '勘定科目コード'
                sheet['M1'].value = '勘定科目'
                sheet['N1'].value = '計画単価'
                sheet['O1'].value = '計画数'
                sheet['P1'].value = '計画合計金額'
                sheet['Q1'].value = '見積単価'
                sheet['R1'].value = '見積数'
                sheet['S1'].value = '見積合計金額'
                sheet['T1'].value = '見積書リンク'
                sheet['U1'].value = '稟議書単価'
                sheet['V1'].value = '稟議書数'
                sheet['W1'].value = '稟議書合計'
                sheet['X1'].value = '稟議書No.'
                sheet['Y1'].value = '稟議書リンク'
                sheet['Z1'].value = 'ワークフローNo.'
                sheet['AA1'].value = '契約金額'
                sheet['AB1'].value = '契約書No.'
                sheet['AC1'].value = '契約書リンク'
                sheet['AD1'].value = '押捺稟議No.'
                sheet['AE1'].value = '押捺稟議リンク'
                sheet['AF1'].value = '注文単価'
                sheet['AG1'].value = '注文数'
                sheet['AH1'].value = '注文合計金額'
                sheet['AI1'].value = '注文書リンク'
                sheet['AJ1'].value = '納品日期限'
                sheet['AK1'].value = '支払日期限'
                sheet['AL1'].value = '納品単価'
                sheet['AM1'].value = '納品数'
                sheet['AN1'].value = '納品合計金額'
                sheet['AO1'].value = '納品書リンク'
                sheet['AP1'].value = '支払単価'
                sheet['AQ1'].value = '購入数'
                sheet['AR1'].value = '購入合計金額'
                sheet['AS1'].value = '請求書リンク情報'
                sheet['AT1'].value = '計画－実績'
                sheet['AU1'].value = '支払繰返し期限'
                sheet['AV1'].value = '購入会社'
                sheet['AW1'].value = '最終データ更新日'
                sheet['AX1'].value = '最終更新者'
                sheet['AY1'].value = '社員番号'
                sheet['AZ1'].value = '担当者名'
                sheet['BA1'].value = 'コメント'
                sheet['BB1'].value = '表示順'

                db = sqlite3.connect('db.sqlite3')
                c = db.cursor()
                c.execute('SELECT * FROM ankenListTable order by hyojijun')
                for i, row in enumerate (c):
                    sheet['A' + str(i+2)].value = row[0]
                    sheet['B' + str(i+2)].value = row[1]
                    sheet['C' + str(i+2)].value = row[2]
                    sheet['D' + str(i+2)].value = row[3]
                    sheet['E' + str(i+2)].value = row[4]
                    sheet['F' + str(i+2)].value = row[5]
                    sheet['G' + str(i+2)].value = row[6]
                    sheet['H' + str(i+2)].value = row[7]
                    sheet['I' + str(i+2)].value = row[8]
                    sheet['J' + str(i+2)].value = row[9]
                    sheet['K' + str(i+2)].value = row[10]
                    sheet['L' + str(i+2)].value = row[11]
                    sheet['M' + str(i+2)].value = row[12]
                    sheet['N' + str(i+2)].value = row[13]
                    sheet['O' + str(i+2)].value = row[14]
                    sheet['P' + str(i+2)].value = row[15]
                    sheet['Q' + str(i+2)].value = row[16]
                    sheet['R' + str(i+2)].value = row[17]
                    sheet['S' + str(i+2)].value = row[18]
                    sheet['T' + str(i+2)].value = row[19]
                    sheet['U' + str(i+2)].value = row[20]
                    sheet['V' + str(i+2)].value = row[21]
                    sheet['W' + str(i+2)].value = row[22]
                    sheet['X' + str(i+2)].value = row[23]
                    sheet['Y' + str(i+2)].value = row[24]
                    sheet['Z' + str(i+2)].value = row[25]
                    sheet['AA' + str(i+2)].value = row[26]
                    sheet['AB' + str(i+2)].value = row[27]
                    sheet['AC' + str(i+2)].value = row[28]
                    sheet['AD' + str(i+2)].value = row[29]
                    sheet['AE' + str(i+2)].value = row[30]
                    sheet['AF' + str(i+2)].value = row[31]
                    sheet['AG' + str(i+2)].value = row[32]
                    sheet['AH' + str(i+2)].value = row[33]
                    sheet['AI' + str(i+2)].value = row[34]
                    sheet['AJ' + str(i+2)].value = row[35]
                    sheet['AK' + str(i+2)].value = row[36]
                    sheet['AL' + str(i+2)].value = row[37]
                    sheet['AM' + str(i+2)].value = row[38]
                    sheet['AN' + str(i+2)].value = row[39]
                    sheet['AO' + str(i+2)].value = row[40]
                    sheet['AP' + str(i+2)].value = row[41]
                    sheet['AQ' + str(i+2)].value = row[42]
                    sheet['AR' + str(i+2)].value = row[43]
                    sheet['AS' + str(i+2)].value = row[44]
                    sheet['AT' + str(i+2)].value = row[45]
                    sheet['AU' + str(i+2)].value = row[46]
                    sheet['AV' + str(i+2)].value = row[47]
                    sheet['AW' + str(i+2)].value = row[48]
                    sheet['AX' + str(i+2)].value = row[49]
                    sheet['AY' + str(i+2)].value = row[50]
                    sheet['AZ' + str(i+2)].value = row[51]
                    sheet['BA' + str(i+2)].value = row[52]
                    sheet['BB' + str(i+2)].value = row[53]

                c.close()
                desktopDir = os.path.expanduser('~/Desktop')
                workbook.save(desktopDir + "/" + inputFile) 

            if fileShurui=="案件ステータスデータ":
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'].value = 'id'
                sheet['B1'].value = 'コード'
                sheet['C1'].value = '案件ステータス'

                db = sqlite3.connect('db.sqlite3')
                c = db.cursor()
                c.execute('SELECT * FROM ankenStatusTable')
                for i, row in enumerate (c):
                    sheet['A' + str(i+2)].value = row[0]
                    sheet['B' + str(i+2)].value = row[1]
                    sheet['C' + str(i+2)].value = row[2]

                c.close()
                desktopDir = os.path.expanduser('~/Desktop')
                workbook.save(desktopDir + "/" + inputFile) 


            if fileShurui=="支払パターンデータ":
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'].value = 'id'
                sheet['B1'].value = 'コード'
                sheet['C1'].value = '支払パターン'

                db = sqlite3.connect('db.sqlite3')
                c = db.cursor()
                c.execute('SELECT * FROM ankenShiharaiPatternTable')
                for i, row in enumerate (c):
                    sheet['A' + str(i+2)].value = row[0]
                    sheet['B' + str(i+2)].value = row[1]
                    sheet['C' + str(i+2)].value = row[2]

                c.close()
                desktopDir = os.path.expanduser('~/Desktop')
                workbook.save(desktopDir + "/" + inputFile) 
 

            if fileShurui=="取引先データ":
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'].value = 'id'
                sheet['B1'].value = 'コード'
                sheet['C1'].value = '会社'
                sheet['D1'].value = '担当者'
                sheet['E1'].value = '住所'
                sheet['F1'].value = '電話'


                db = sqlite3.connect('db.sqlite3')
                c = db.cursor()
                c.execute('SELECT * FROM ankenTorihikisakiTable')
                for i, row in enumerate (c):
                    sheet['A' + str(i+2)].value = row[0]
                    sheet['B' + str(i+2)].value = row[1]
                    sheet['C' + str(i+2)].value = row[2]
                    sheet['D' + str(i+2)].value = row[3]
                    sheet['E' + str(i+2)].value = row[4]
                    sheet['F' + str(i+2)].value = row[5]

                c.close()
                desktopDir = os.path.expanduser('~/Desktop')
                workbook.save(desktopDir + "/" + inputFile) 


            if fileShurui=="担当者データ":
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'].value = 'id'
                sheet['B1'].value = 'コード'
                sheet['C1'].value = '社員名'


                db = sqlite3.connect('db.sqlite3')
                c = db.cursor()
                c.execute('SELECT * FROM ankenTantoshaTable')
                for i, row in enumerate (c):
                    sheet['A' + str(i+2)].value = row[0]
                    sheet['B' + str(i+2)].value = row[1]
                    sheet['C' + str(i+2)].value = row[2]

                c.close()
                desktopDir = os.path.expanduser('~/Desktop')
                workbook.save(desktopDir + "/" + inputFile) 


            if fileShurui=="勘定科目データ":
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet['A1'].value = 'id'
                sheet['B1'].value = 'コード'
                sheet['C1'].value = '勘定科目'


                db = sqlite3.connect('db.sqlite3')
                c = db.cursor()
                c.execute('SELECT * FROM ankenKanjokamokuTable')
                for i, row in enumerate (c):
                    sheet['A' + str(i+2)].value = row[0]
                    sheet['B' + str(i+2)].value = row[1]
                    sheet['C' + str(i+2)].value = row[2]

                c.close()
                desktopDir = os.path.expanduser('~/Desktop')
                workbook.save(desktopDir + "/" + inputFile) 


        
    return render(request,'ankenkanri2/export2.html',context)


        
def dataimport (request):
    kazu = AnkenList.objects.all().count()
    context = {
        'ankenList' : AnkenList.objects.order_by('hyojijun'),
        'kazu' : kazu
        
    }
    return render(request,'ankenkanri2/import.html',context)
   
   
def dataexport (request):
    kazu = AnkenList.objects.all().count()
    
    context = {
        'ankenList' : AnkenList.objects.order_by('hyojijun'),
        'kazu' : kazu
    }
    return render(request,'ankenkanri2/export.html',context)
     

    
def import2 (request):
    os.environ.setdefault('DJANGO_SETTINGS_MODULE','first_project.settings')
    from django import setup
    setup()
    kazu = AnkenList.objects.all().count()
    inputFile=request.POST.get('inputFile')
    desktopDir = os.path.expanduser('~/Desktop')
    fileShurui=request.POST.get('tableName')
    message=[]
    # hyojibango=[]
    # komokumei=[]
    # message=[]
    
    
    # print('インプットファイルは',inputFile)
    if request.method == "POST":
        if "inputFile" in request.POST:
            
            workbook = openpyxl.load_workbook(desktopDir + "/" + inputFile)
            sheet = workbook.active
            it = sheet.iter_rows(min_row=2)
                #エラーメッセージ用
            if fileShurui == "案件管理データ":    
                for row in it:
                            
                    #必須項目の確認
                    if row[1].value is None:
                        message.append(str(row[53].value)+"行目の【経理承認】が空白です")
                    if row[2].value is None:
                        message.append(str(row[53].value)+"行目の【案件ステータスコード】が空白です")
                    if row[3].value is None:
                        message.append(str(row[53].value)+"行目の【案件ステータス】が空白です")
                    if row[5].value is None:
                        message.append(str(row[53].value)+"行目の【支払パターン】が空白です")
                    if row[6].value is None:
                        message.append(str(row[53].value)+"行目の【管理番号】が空白です")
                    if row[7].value is None:
                        message.append(str(row[53].value)+"行目の【枝番】が空白です")
                    if row[8].value is None:
                        message.append(str(row[53].valu)+"行目の【案件名】が空白です")
                    if row[9].value is None:
                        message.append(str(row[53].value)+"行目の【取引先コード】が空白です")
                    if row[10].value is None:
                        message.append(str(row[53].value)+"行目の【取引先名】が空白です")
                    if row[13].value is None:
                        message.append(str(row[53].value)+"行目の【計画単価】が空白です")
                    if row[14].value is None:
                        message.append(str(row[53].value)+"行目の【計画数】が空白です")
                    
                    # #「請求書入手済」ステータスにおける追加必須項目
                    if row[2].value == 9 and row[41].value is None:
                        message.append(str(row[53].value)+"行目の【支払単価】が空白")
                    if row[2].value == 9 and row[42].value is None:
                        message.append(str(row[53].value)+"行目の【購入数】が空白です")
                    if row[2].value == 9 and row[43].value is None:
                        message.append(str(row[53].value)+"行目の【購入合計】が空白です")
                        
                    #型の確認
                    # if re.compile(r'[12]\d{3}[/\-年](0?[1-9]|1[0-2])[/\-月]([12][0-9]|3[01]|0?[1-9])日?').match(str(row[35].value)):
                    #     pass
                    # else:
                    #     hyojibango.append(row[53].value)
                    #     komokumei.append("納品期限")
                    #     message.append("正しい日付ではない")

                    # if re.compile(r'[12]\d{3}[/\-年](0?[1-9]|1[0-2])[/\-月]([12][0-9]|3[01]|0?[1-9])日?').match(str(row[36].value)):
                    #     pass
                    # else:
                    #     hyojibango.append(row[53].value)
                    #     komokumei.append("支払期限")
                    #     message.append("正しい日付ではない")

                    # if re.compile(r'[12]\d{3}[/\-年](0?[1-9]|1[0-2])[/\-月]([12][0-9]|3[01]|0?[1-9])日?').match(str(row[46].value)):
                    #     pass
                    # else:
                    #     hyojibango.append(row[53].value)
                    #     komokumei.append("繰り返し期限")
                    #     message.append("正しい日付ではない")
                
                if len(message)!=0:
                    pass
                
                else:
                    kanriNos = []
                    edabans = []
                    it = sheet.iter_rows(min_row=2)
                    for row in it:
                        kanriNos.append(row[6].value)
                        edabans.append(row[7].value)
            
                        # print('IDの値は',row[0].value)
                        # AnkenList.objects.update_or_create (
                        ankn, created=AnkenList.objects.update_or_create (    
                            kanriNo= row[6].value,
                            edaban= row[7].value,
                            defaults={
                                "keiriShonin" : row[1].value,
                                "statusCode" : row[2].value,
                                "status" : row[3].value,
                                "edabanGroup" : row[4].value,
                                "shiharaiPattern" : row[5].value,
                                "kanriNo" : row[6].value,
                                "edaban" : row[7].value,
                                "ankenMei" : row[8].value,
                                "torihikisakiCode" : row[9].value,
                                "torihikisakiMei" : row[10].value,
                                "kanjokamokuCode" : row[11].value,
                                "kanjokamoku" : row[12].value,
                                "keikakuTanka" : row[13].value,
                                "keikakuSuu" : row[14].value,
                                "keikakuGokei" : row[15].value,
                                "mitsumoriTanka" : row[16].value,
                                "mitsumoriSuu" : row[17].value,
                                "mitsumoriGokei" : row[18].value,
                                "mitsumoriLink" : row[19].value,
                                "ringiTanka" : row[20].value,
                                "ringiSuu" : row[21].value,
                                "ringiGokei" : row[22].value,
                                "ringishoNo" : row[23].value,
                                "ringishoLink" : row[24].value,
                                "wfNo" : row[25].value,
                                "keiyakuKingaku" : row[26].value,
                                "keiyakushoNo" : row[27].value,
                                "keiyakushoLink" : row[28].value,
                                "onatsuRingiNo" : row[29].value,
                                "onatsuRingiLink" : row[30].value,
                                "chumonTanka" : row[31].value,
                                "chumonSuu" : row[32].value,
                                "chumonGokei" : row[33].value,
                                "chumonLink" : row[34].value,
                                "nohinKigen" : row[35].value,
                                "shiharaiKigen" : row[36].value,
                                "nohinTanka" : row[37].value,
                                "nohinSuu" : row[38].value,
                                "nohinGokei" : row[39].value,
                                "nohinLink" : row[40].value,
                                "shiharaiTanka" : row[41].value,
                                "konyuSuu" : row[42].value,
                                "konyuGokei" : row[43].value,
                                "seikyushoLink" : row[44].value,
                                "keikaku_Jisseki" : row[45].value,
                                "kurikaeshiKigen" : row[46].value,
                                "konyuKaisha" : row[47].value,
                                "dataKoshinbi" : row[48].value,
                                "saishuKoshinsha" : row[49].value,
                                "shainNo" : row[50].value,
                                "tantosha" : row[51].value,
                                "comment" : row[52].value,
                                "hyojijun" : row[53].value,

                            }    
                        )
                        

                    # print('ＩＤリスト：',ids) 
                    ankenList=AnkenList.objects.order_by('hyojijun')   
                    for anken in ankenList:
                        hantei=0
                        
                        for i in range(len(kanriNos)):
                            if anken.kanriNo==kanriNos[i]:
                                if anken.edaban==edabans[i]:
                                    hantei = 1
                        if hantei==0:
                            AnkenList.objects.filter(id=anken.id).delete()                              
                i = 1
                anken = AnkenList.objects.order_by('hyojijun')
                for ankens in anken:
                    ankens.hyojijun = i
                    ankens.save()
                    i=i+1
            
            if fileShurui == "案件ステータスデータ":    
                it = sheet.iter_rows(min_row=2)
                for row in it:
                    ankn, created=AnkenStatus.objects.update_or_create (    
                        id= row[0].value,
                        defaults={
                            "ankenStatusCode" : row[1].value,
                            "ankenStatus" : row[2].value,
                        }    
                    )

            if fileShurui == "支払パターンデータ":    
                it = sheet.iter_rows(min_row=2)
                for row in it:
                    ankn, created=AnkenShiharaiPattern.objects.update_or_create (    
                        id= row[0].value,
                        defaults={
                            "ankenShiharaiPatternCode" : row[1].value,
                            "abkenShiharaiPattern" : row[2].value,
                        }    
                    )


            if fileShurui == "取引先データ":    
                it = sheet.iter_rows(min_row=2)
                for row in it:
                    ankn, created=AnkenTorihikisaki.objects.update_or_create (    
                        id= row[0].value,
                        defaults={
                            "ankenTorihikisakiCode" : row[1].value,
                            "ankenTorihikisakiKaisha" : row[2].value,
                            "ankenTorihikisakiTantosha" : row[3].value,
                            "ankenTorihikisakiJusho" : row[4].value,
                            "ankenTorihikisakiTel" : row[5].value,
                        }    
                    )

            if fileShurui == "担当者データ":    
                it = sheet.iter_rows(min_row=2)
                for row in it:
                    ankn, created=AnkenTantosha.objects.update_or_create (    
                        id= row[0].value,
                        defaults={
                            "ankenTantoshaCode" : row[1].value,
                            "ankenTantoshaMei" : row[2].value,
                        }    
                    )


            if fileShurui == "勘定科目データ":    
                it = sheet.iter_rows(min_row=2)
                for row in it:
                    ankn, created=AnkenKanjokamoku.objects.update_or_create (    
                        id= row[0].value,
                        defaults={
                            "ankenKanjokamokuCode" : row[1].value,
                            "ankenKanjokamoku" : row[2].value,
                        }    
                    )


    context = {
        'ankenList' : AnkenList.objects.order_by('hyojijun'),
        'inputFile' : inputFile,
        'kazu' : kazu,
        'message' : message,
        'shurui' : fileShurui
        
    }            
             
    # print("メッセージは",message)      
    return render(request,'ankenkanri2/import2.html',context)


def importTest (request):
    os.environ.setdefault('DJANGO_SETTINGS_MODULE','first_project.settings')
    from django import setup
    setup()
    # AnkenList.objects.create(
    a = AnkenList(
        
        keiriShonin= True,
        statusCord= 0,
        status= '案件入力（作成中）',
        edabanGroup= 3,
        shiharaiPattern= 3,
        kanriNo= 'S22-010',
        edaban= '1',
        ankenMei= 'Office365 月額費用　4月',
        saishuKoshinsha= '荒木',
        shainNo= 17059,
        tantosha= '荒木',
        comment= '・・・・'
    )
    a.save()
    context = {
        'ankenList' : AnkenList.objects.all(),
        
    }
    return render(request,'ankenkanri2/import2.html',context)


## TOPメニュー表示中に編集ボタンが押された
def edit(request):
    form = forms.formAnkenList()
    kazu = AnkenList.objects.all().count()    
    return render(
        request,'ankenkanri2/edit.html',context={'form':form,'ankenList' \
                                              : AnkenList.objects.order_by('hyojijun'),'kazu':kazu}
    )


## 編集画面表示中にボタンが押された
def editOperation(request):
    if "returnTopButton" in request.POST:
        request.session['sessionDisplayCode'] = 'dp00' 
        #return redirect('/ankenkanri2/index')
        kazu = AnkenList.objects.all().count()
        context = {
            'ankenList' : AnkenList.objects.order_by('hyojijun'),
            'kazu' : kazu
        }
        request.session['sessionDisplayCode'] = 'dp00' 
        return render(request,'ankenkanri2/index.html',context)
    else:
        #context = contextFormSet(request)
        kazu = AnkenList.objects.all().count()
        checkType = ''
        checkBoxValue=''
        if request.method == 'POST':
            form = forms.formAnkenList()
            kanrishori=request.POST.getlist('kanriShori')
            henko=request.POST.getlist('henko')
            hukusha=request.POST.getlist('hukusha')
            sakujo=request.POST.getlist('sakujo')
            # print("えでぃっと2でのポストしてきた管理処理は",kanrishori)
            # print("えでぃっと2でのポストしてきた変更は",henko)
            # print("えでぃっと2でのポストしてきた複写は",hukusha)
            # print("えでぃっと2でのポストしてきた削除は",sakujo)
            
            obj = AnkenList.objects.order_by('hyojijun')
            if kanrishori:
                obj=AnkenList.objects.filter(id__in=list(map(int, kanrishori)))
                checkBoxValue = kanrishori
                checkType = 'kanrishori'
            elif henko:
                obj=AnkenList.objects.filter(id__in=list(map(int, henko)))
                checkBoxValue = henko
                checkType = 'henko'    
            elif hukusha:
                obj=AnkenList.objects.filter(id__in=list(map(int, hukusha)))
                checkBoxValue = hukusha 
                checkType = 'hukusha'   
            elif sakujo:
                obj=AnkenList.objects.filter(id__in=list(map(int, sakujo)))
                checkBoxValue = sakujo    
                checkType = 'sakujo' 
            # else:
            #     # obj=AnkenList.objects.filter(id__in=list(map(int, sakujo)))
            #     checkBoxValue = None    
            #     checkType = 'nashi' 
                
            
            initialValues={
                "id":obj[0].id,
                "keiriShonin":obj[0].keiriShonin,
                "statusCode":obj[0].statusCode,
                "status":obj[0].status,
                "edabanGroup":obj[0].edabanGroup,
                "shiharaiPattern":obj[0].shiharaiPattern,
                "kanriNo":obj[0].kanriNo,
                "edaban":obj[0].edaban,
                "ankenMei":obj[0].ankenMei,
                "torihikisakiCode":obj[0].torihikisakiCode,
                "torihikisakiMei":obj[0].torihikisakiMei,
                "kanjokamokuCode":obj[0].kanjokamokuCode,
                "kanjokamoku":obj[0].kanjokamoku,
                "keikakuTanka":obj[0].keikakuTanka,
                "keikakuSuu":obj[0].keikakuSuu,
                "keikakuGokei":obj[0].keikakuGokei,
                "mitsumoriTanka":obj[0].mitsumoriTanka,
                "mitsumoriSuu":obj[0].mitsumoriSuu,
                "mitsumoriGokei":obj[0].mitsumoriGokei,
                "mitsumoriLink":obj[0].mitsumoriLink,
                "ringiTanka":obj[0].ringiTanka,
                "ringiSuu":obj[0].ringiSuu,
                "ringiGokei":obj[0].ringiGokei,
                "ringishoNo":obj[0].ringishoNo,
                "ringishoLink":obj[0].ringishoLink,
                "wfNo":obj[0].wfNo,
                "keiyakuKingaku":obj[0].keiyakuKingaku,
                "keiyakushoNo":obj[0].keiyakushoNo,
                "keiyakushoLink":obj[0].keiyakushoLink,
                "onatsuRingiNo":obj[0].onatsuRingiNo,
                "onatsuRingiLink":obj[0].onatsuRingiLink,
                "chumonTanka":obj[0].chumonTanka,
                "chumonSuu":obj[0].chumonSuu,
                "chumonGokei":obj[0].chumonGokei,
                "chumonLink":obj[0].chumonLink,
                "nohinKigen":obj[0].nohinKigen,
                "shiharaiKigen":obj[0].shiharaiKigen,
                "nohinTanka":obj[0].nohinTanka,
                "nohinSuu":obj[0].nohinSuu,
                "nohinGokei":obj[0].nohinGokei,
                "nohinLink":obj[0].nohinLink,
                "shiharaiTanka":obj[0].shiharaiTanka,
                "konyuSuu":obj[0].konyuSuu,
                "konyuGokei":obj[0].konyuGokei,
                "seikyushoLink":obj[0].seikyushoLink,
                "keikaku_Jisseki":obj[0].keikaku_Jisseki,
                "kurikaeshiKigen":obj[0].kurikaeshiKigen,
                "konyuKaisha":obj[0].konyuKaisha,
                "dataKoshinbi":obj[0].dataKoshinbi,
                "saishuKoshinsha":obj[0].saishuKoshinsha,
                "shainNo":obj[0].shainNo,
                "tantosha":obj[0].tantosha,
                "comment":obj[0].comment,
                "hyojijun":obj[0].hyojijun
            }
            form=forms.formAnkenList(initialValues or request.POST)
        # print("えでぃっと2でのちぇっくたいぷは",checkType)
        return render(
            request,'ankenkanri2/edit2.html', context={'checkType':checkType,'form':form,'ankenList' \
                        : AnkenList.objects.filter(id__in=list(map(int, checkBoxValue))),'kazu':kazu}
        )

        request.session['sessionDisplayCode'] = 'dp231'   
        return render('ankenkanri2/edit2.html',context)



def contextFormSet(request):
    kazu = AnkenList.objects.all().count()
    if request.method == 'POST':
        form = forms.formAnkenList()
        kanrishori=request.POST.getlist('kanriShori')
        henko=request.POST.getlist('henko')
        hukusha=request.POST.getlist('hukusha')
        sakujo=request.POST.getlist('sakujo')
        
        obj = AnkenList.objects.order_by('hyojijun')
        checkType = ''
        checkBoxValue=''
        if kanrishori:
            obj=AnkenList.objects.filter(id__in=list(map(int, kanrishori)))
            checkBoxValue = kanrishori
            checkType = 'kanrishori'
        elif henko:
            obj=AnkenList.objects.filter(id__in=list(map(int, henko)))
            checkBoxValue = henko
            checkType = 'henko'    
        elif hukusha:
            obj=AnkenList.objects.filter(id__in=list(map(int, hukusha)))
            checkBoxValue = hukusha 
            checkType = 'hukusha'   
        elif sakujo:
            obj=AnkenList.objects.filter(id__in=list(map(int, sakujo)))
            checkBoxValue = sakujo    
            checkType = 'sakujo' 
                    
        initialValues={
            "id":obj[0].id,
            "keiriShonin":obj[0].keiriShonin,
            "statusCode":obj[0].statusCode,
            "status":obj[0].status,
            "edabanGroup":obj[0].edabanGroup,
            "shiharaiPattern":obj[0].shiharaiPattern,
            "kanriNo":obj[0].kanriNo,
            "edaban":obj[0].edaban,
            "ankenMei":obj[0].ankenMei,
            "torihikisakiCode":obj[0].torihikisakiCode,
            "torihikisakiMei":obj[0].torihikisakiMei,
            "kanjokamokuCode":obj[0].kanjokamokuCode,
            "kanjokamoku":obj[0].kanjokamoku,
            "keikakuTanka":obj[0].keikakuTanka,
            "keikakuSuu":obj[0].keikakuSuu,
            "keikakuGokei":obj[0].keikakuGokei,
            "mitsumoriTanka":obj[0].mitsumoriTanka,
            "mitsumoriSuu":obj[0].mitsumoriSuu,
            "mitsumoriGokei":obj[0].mitsumoriGokei,
            "mitsumoriLink":obj[0].mitsumoriLink,
            "ringiTanka":obj[0].ringiTanka,
            "ringiSuu":obj[0].ringiSuu,
            "ringiGokei":obj[0].ringiGokei,
            "ringishoNo":obj[0].ringishoNo,
            "ringishoLink":obj[0].ringishoLink,
            "wfNo":obj[0].wfNo,
            "keiyakuKingaku":obj[0].keiyakuKingaku,
            "keiyakushoNo":obj[0].keiyakushoNo,
            "keiyakushoLink":obj[0].keiyakushoLink,
            "onatsuRingiNo":obj[0].onatsuRingiNo,
            "onatsuRingiLink":obj[0].onatsuRingiLink,
            "chumonTanka":obj[0].chumonTanka,
            "chumonSuu":obj[0].chumonSuu,
            "chumonGokei":obj[0].chumonGokei,
            "chumonLink":obj[0].chumonLink,
            "nohinKigen":obj[0].nohinKigen,
            "shiharaiKigen":obj[0].shiharaiKigen,
            "nohinTanka":obj[0].nohinTanka,
            "nohinSuu":obj[0].nohinSuu,
            "nohinGokei":obj[0].nohinGokei,
            "nohinLink":obj[0].nohinLink,
            "shiharaiTanka":obj[0].shiharaiTanka,
            "konyuSuu":obj[0].konyuSuu,
            "konyuGokei":obj[0].konyuGokei,
            "seikyushoLink":obj[0].seikyushoLink,
            "keikaku_Jisseki":obj[0].keikaku_Jisseki,
            "kurikaeshiKigen":obj[0].kurikaeshiKigen,
            "konyuKaisha":obj[0].konyuKaisha,
            "dataKoshinbi":obj[0].dataKoshinbi,
            "saishuKoshinsha":obj[0].saishuKoshinsha,
            "shainNo":obj[0].shainNo,
            "tantosha":obj[0].tantosha,
            "comment":obj[0].comment,
            "hyojijun":obj[0].hyojijun
        }
        form=forms.formAnkenList(initialValues or request.POST)
        context = {'checkType':checkType,'form':form,'ankenList' \
                    : AnkenList.objects.filter(id__in=list(map(int, checkBoxValue))),'kazu':kazu}
        return context

def edit2(request):
    # print("りくえすとめそっどは",request.method)
#    context = contextFormSet(request)
    kazu = AnkenList.objects.all().count()
    checkType = ''
    checkBoxValue=''
    if request.method == 'POST':
        form = forms.formAnkenList()
        kanrishori=request.POST.getlist('kanriShori')
        henko=request.POST.getlist('henko')
        hukusha=request.POST.getlist('hukusha')
        sakujo=request.POST.getlist('sakujo')
        # print("えでぃっと2でのポストしてきた管理処理は",kanrishori)
        # print("えでぃっと2でのポストしてきた変更は",henko)
        # print("えでぃっと2でのポストしてきた複写は",hukusha)
        # print("えでぃっと2でのポストしてきた削除は",sakujo)
        
        obj = AnkenList.objects.order_by('hyojijun')
        if kanrishori:
            obj=AnkenList.objects.filter(id__in=list(map(int, kanrishori)))
            checkBoxValue = kanrishori
            checkType = 'kanrishori'
        elif henko:
            obj=AnkenList.objects.filter(id__in=list(map(int, henko)))
            checkBoxValue = henko
            checkType = 'henko'    
        elif hukusha:
            obj=AnkenList.objects.filter(id__in=list(map(int, hukusha)))
            checkBoxValue = hukusha 
            checkType = 'hukusha'   
        elif sakujo:
            obj=AnkenList.objects.filter(id__in=list(map(int, sakujo)))
            checkBoxValue = sakujo    
            checkType = 'sakujo' 
        # else:
        #     # obj=AnkenList.objects.filter(id__in=list(map(int, sakujo)))
        #     checkBoxValue = None    
        #     checkType = 'nashi' 
            
        
        initialValues={
            "id":obj[0].id,
            "keiriShonin":obj[0].keiriShonin,
            "statusCode":obj[0].statusCode,
            "status":obj[0].status,
            "edabanGroup":obj[0].edabanGroup,
            "shiharaiPattern":obj[0].shiharaiPattern,
            "kanriNo":obj[0].kanriNo,
            "edaban":obj[0].edaban,
            "ankenMei":obj[0].ankenMei,
            "torihikisakiCode":obj[0].torihikisakiCode,
            "torihikisakiMei":obj[0].torihikisakiMei,
            "kanjokamokuCode":obj[0].kanjokamokuCode,
            "kanjokamoku":obj[0].kanjokamoku,
            "keikakuTanka":obj[0].keikakuTanka,
            "keikakuSuu":obj[0].keikakuSuu,
            "keikakuGokei":obj[0].keikakuGokei,
            "mitsumoriTanka":obj[0].mitsumoriTanka,
            "mitsumoriSuu":obj[0].mitsumoriSuu,
            "mitsumoriGokei":obj[0].mitsumoriGokei,
            "mitsumoriLink":obj[0].mitsumoriLink,
            "ringiTanka":obj[0].ringiTanka,
            "ringiSuu":obj[0].ringiSuu,
            "ringiGokei":obj[0].ringiGokei,
            "ringishoNo":obj[0].ringishoNo,
            "ringishoLink":obj[0].ringishoLink,
            "wfNo":obj[0].wfNo,
            "keiyakuKingaku":obj[0].keiyakuKingaku,
            "keiyakushoNo":obj[0].keiyakushoNo,
            "keiyakushoLink":obj[0].keiyakushoLink,
            "onatsuRingiNo":obj[0].onatsuRingiNo,
            "onatsuRingiLink":obj[0].onatsuRingiLink,
            "chumonTanka":obj[0].chumonTanka,
            "chumonSuu":obj[0].chumonSuu,
            "chumonGokei":obj[0].chumonGokei,
            "chumonLink":obj[0].chumonLink,
            "nohinKigen":obj[0].nohinKigen,
            "shiharaiKigen":obj[0].shiharaiKigen,
            "nohinTanka":obj[0].nohinTanka,
            "nohinSuu":obj[0].nohinSuu,
            "nohinGokei":obj[0].nohinGokei,
            "nohinLink":obj[0].nohinLink,
            "shiharaiTanka":obj[0].shiharaiTanka,
            "konyuSuu":obj[0].konyuSuu,
            "konyuGokei":obj[0].konyuGokei,
            "seikyushoLink":obj[0].seikyushoLink,
            "keikaku_Jisseki":obj[0].keikaku_Jisseki,
            "kurikaeshiKigen":obj[0].kurikaeshiKigen,
            "konyuKaisha":obj[0].konyuKaisha,
            "dataKoshinbi":obj[0].dataKoshinbi,
            "saishuKoshinsha":obj[0].saishuKoshinsha,
            "shainNo":obj[0].shainNo,
            "tantosha":obj[0].tantosha,
            "comment":obj[0].comment,
            "hyojijun":obj[0].hyojijun
        }
        form=forms.formAnkenList(initialValues or request.POST)
    # print("えでぃっと2でのちぇっくたいぷは",checkType)
    return render(
        request,'ankenkanri2/edit2.html', context={'checkType':checkType,'form':form,'ankenList' \
                    : AnkenList.objects.filter(id__in=list(map(int, checkBoxValue))),'kazu':kazu}
    )

    
def edit3(request):
    # from .views1 import contextSet
    
    kazu = AnkenList.objects.all().count()
    os.environ.setdefault('DJANGO_SETTINGS_MODULE','first_project.settings')
    from django import setup
    setup()
    form = forms.formAnkenList()
    message =[]
    # print( "りくえすとめそっどは",request.method)
    if request.method == 'POST':
        form = forms.formAnkenList(request.POST)
        checkType = request.POST['checkType']
        # print("さぶみっとぼたんは",request.POST['checkType'])
        
        if checkType == "kanrishori":
            if form.is_valid():
                inData = form.cleaned_data.get('kanriNo')
                edaban = form.cleaned_data.get('edaban')
                
                # context = contextSet(inData,edaban)
                
                model = AnkenList
                ctx = {}
                if (edaban is None) or (edaban==""):
                    q = AnkenList.objects.filter(kanriNo=inData).filter(edaban__isnull=True)
                else:
                    q = AnkenList.objects.filter(kanriNo=inData).filter(edaban=edaban)

                ctx["object"] = q
                
                
                badgeName = ["案件入力済\n(STATUS = 1)", \
                            "見積書依頼処理\n(STATUS = 2)", \
                            "見積書入手処理\n(STATUS = 3)", \
                            "稟議承認処理\n(STATUS = 4)", \
                            "契約書_作成完了\n(STATUS = 5)",\
                            "契約書_締結完了\n(STATUS = 6)", \
                            "注文処理\n(STATUS = 7)", \
                            "納品処理\n(STATUS = 8)", \
                            "請求書処理\n(STATUS = 9)", \
                            "支払処理済\n(STATUS = 10)"
                            ]
                buttonName0 = {"ankenNyuuryokuKan1"     : "案件入力を終了しました", \
                            "mitsumorisyoIraiKan2"   : "見積書の作成依頼を終了しました", \
                            "mitumorisyoNyuusyuKan3" : "見積書を入手しました", \
                            "ringiShoninKan4"        : "稟議決裁を終了しました", \
                            "keiyakusyoSakuseiKan5"  : "契約書の作成が終了しました", \
                            "keiyakusyoTeiketsuKan6" : "契約書の締結が終了しました", \
                            "cyuumonKan7"            : "注文処理が終了しました", \
                            "nouhinKan8"             : "納品処理が終了しました", \
                            "seikyuusyoNyuusyu9"     : "請求書処理が終了しました", \
                            "shiharaiKan10"          : "支払処理が終了しました"
                            }

                buttonName = ["案件入力を終了しました", "見積書の作成依頼を終了しました", \
                            "見積書を入手しました", "稟議決裁を終了しました", \
                            "契約書の作成が終了しました", "契約書の締結が終了しました", \
                            "注文処理が終了しました", "納品処理が終了しました", \
                            "請求書処理が終了しました", "支払処理が終了しました"
                            ]
                buttonName2 = ["ankenNyuuryokuKan1", "mitsumorisyoIraiKan2", \
                            "mitumorisyoNyuusyuKan3", "ringiShoninKan4", \
                            "keiyakusyoSakuseiKan5", "keiyakusyoTeiketsuKan6", \
                            "cyuumonKan7", "nouhinKan8", \
                            "seikyuusyoNyuusyu9", "shiharaiKan10" 
                            ]                
                
               
                print('(contextSet)q[0] =', q[0])
                badgeGray = q[0].statusCode
                buttonGray = q[0].statusCode
                context = {
                    'badgeName': badgeName,
                    'buttonName0': buttonName0,
                    'buttonName': buttonName,
                    'buttonName2': buttonName2,
                    'badgeGray': badgeGray,
                    'buttonGray': buttonGray,
                    'q': q,
                }
                
                
                # page_n(request)
                # request.session['sessionDisplayCode'] = 'dp02'
                # request.session['sessionKanriNo'] = form.cleaned_data.get('kanriNo')
                # request.session['sessionEdaban'] = form.cleaned_data.get('edaban')
                return render(request, 'ankenkanri/page_n.html',context)
        if checkType == "henko":
            if form.is_valid():
                #入力規制
                #空白判定
                if form.cleaned_data.get('keiriShonin') is None:
                    message.append("【経理承認】が空白です")
                if form.cleaned_data.get('statusCode') is None:
                    message.append("【案件ステータスコード】が空白です")
                if form.cleaned_data.get('shiharaiPattern') is None:
                    message.append("【支払パターン】が空白です")
                if form.cleaned_data.get('kanriNo') is None:
                    message.append("【管理番号】が空白です")
                if form.cleaned_data.get('ankenMei') is None:
                    message.append("【案件名】が空白です")
                if form.cleaned_data.get('torihikisakiCode') is None:
                    message.append("【取引先コード】が空白です")
                if form.cleaned_data.get('keikakuTanka') is None:
                    message.append("【計画単価】が空白です")
                if form.cleaned_data.get('keikakuSuu') is None:
                    message.append("【計画数】が空白です")
                #ステータスによる期日空白判定
                if form.cleaned_data.get('statusCode')==9 and form.cleaned_data.get('shiharaiTanka')is None:
                    message.append("【支払単価】が空白です")
                if form.cleaned_data.get('statusCode')==9 and form.cleaned_data.get('konyuSuu')is None:
                    message.append("【購入数】が空白です")                
                
                if len(message)!=0:
                    pass
            
                else:                
                    anken = AnkenList.objects.order_by('hyojijun')
                    statuscode = form.cleaned_data.get('statusCode')
                    statusanken = AnkenStatus.objects.filter(ankenStatusCode=int(statuscode))
                    status = statusanken[0].ankenStatus
                                    
                    torihikisakicode = form.cleaned_data.get('torihikisakiCode')
                    torihikisakianken = AnkenTorihikisaki.objects.filter(ankenTorihikisakiCode=torihikisakicode)
                    torihikisaki = torihikisakianken[0].ankenTorihikisakiKaisha
                    
                    tantoshacode = form.cleaned_data.get('shainNo')
                    tantoshaanken = AnkenTantosha.objects.filter(ankenTantoshaCode=tantoshacode)
                    tantosha = tantoshaanken[0].ankenTantoshaMei
                    
                    kanjokamokucode = form.cleaned_data.get('kanjokamokuCode')
                    kanjokamokuanken = AnkenKanjokamoku.objects.filter(ankenKanjokamokuCode=kanjokamokucode)
                    kanjokamoku = kanjokamokuanken[0].ankenKanjokamoku
                    
                    keikakugokei=NULL
                    mitsumorigokei=NULL
                    ringigokei  =NULL
                    chumongokei =NULL
                    nohingokei =NULL
                    konyugokei =NULL
                    
                    
                    if form.cleaned_data.get('keikakuTanka')is not None and form.cleaned_data.get('keikakuSuu')is not None:
                        keikakugokei = int(form.cleaned_data.get('keikakuTanka'))*int(form.cleaned_data.get('keikakuSuu'))
                    # if (form.cleaned_data.get('mitsumoriTanka')!=0 or form.cleaned_data.get('mitsumoriTanka') is not None)and (form.cleaned_data.get('mitsumoriSuu')!=0 or form.cleaned_data.get('mitsumoriSuu')is not None):
                    if form.cleaned_data.get('mitsumoriTanka') is not None and form.cleaned_data.get('mitsumoriSuu')is not None:
                        mitsumorigokei = int(form.cleaned_data.get('mitsumoriTanka'))*int(form.cleaned_data.get('mitsumoriSuu'))
                    if form.cleaned_data.get('ringiTanka') is not None and  form.cleaned_data.get('ringiSuu')is not None:                   
                        ringigokei = int(form.cleaned_data.get('ringiTanka'))*int(form.cleaned_data.get('ringiSuu'))
                    if form.cleaned_data.get('chumonTanka')is not None and form.cleaned_data.get('chumonSuu')is not None:
                        chumongokei = int(form.cleaned_data.get('chumonTanka'))*int(form.cleaned_data.get('chumonSuu'))
                    if form.cleaned_data.get('nohinTanka')is not None and form.cleaned_data.get('nohinSuu')is not None:
                        nohingokei = int(form.cleaned_data.get('nohinTanka'))*int(form.cleaned_data.get('nohinSuu'))
                    if form.cleaned_data.get('shiharaiTanka')is not None and form.cleaned_data.get('konyuSuu')is not None:
                        konyugokei = int(form.cleaned_data.get('shiharaiTanka'))*int(form.cleaned_data.get('konyuSuu'))
                                
                    kanriNo=form.cleaned_data.get('kanriNo')
                    edaban=form.cleaned_data.get('edaban')
                    hyojijuns =form.cleaned_data.get('hyojijun')
                    edabanShitei = edaban
                                    
                    for ankens in anken:
                        if ankens.kanriNo==kanriNo:
                            if ankens.edaban > edaban:
                                edabanShitei=int(ankens.edaban) + 1
                    
                    
                    
                    # edabanShitei=0
                    for ankens in anken:
                        
                        # if ankens.kanriNo == kanriNo and int(ankens.edaban) > edabanShitei:
                        #     edabanShitei = int(ankens.edaban)
                        if ankens.hyojijun >=hyojijuns:
                            ankens.hyojijun=ankens.hyojijun+1
                            ankens.save()
                    
                    for ankens in anken:
                        if ankens.hyojijun == hyojijuns:
                            hyojijuns=hyojijuns+1
                    
                    if int(statuscode) == 99:
                        edabanShitei = 0 
                    
                    ankn, created=AnkenList.objects.update_or_create (    
                        kanriNo= form.cleaned_data.get('kanriNo'),
                        edaban= form.cleaned_data.get('edaban'),
                        defaults={
                            "keiriShonin" : form.cleaned_data.get('keiriShonin'),
                            "statusCode" : form.cleaned_data.get('statusCode'),
                            # "status" : form.cleaned_data.get('status'),
                            "status" : status,
                            "edabanGroup" : form.cleaned_data.get('edabanGroup'),
                            "shiharaiPattern" : form.cleaned_data.get('shiharaiPattern'),
                            "kanriNo" : form.cleaned_data.get('kanriNo'),
                            "edaban" : edabanShitei,
                            "ankenMei" : form.cleaned_data.get('ankenMei'),
                            "torihikisakiCode" : form.cleaned_data.get('torihikisakiCode'),
                            # "torihikisakiMei" : form.cleaned_data.get('torihikisakiMei'),
                            "torihikisakiMei" : torihikisaki,
                            "kanjokamokuCode" : form.cleaned_data.get('kanjokamokuCode'),
                            "kanjokamoku" : kanjokamoku,
                            "keikakuTanka" : form.cleaned_data.get('keikakuTanka'),
                            "keikakuSuu" : form.cleaned_data.get('keikakuSuu'),
                            "keikakuGokei" : keikakugokei,
                            "mitsumoriTanka" : form.cleaned_data.get('mitsumoriTanka'),
                            "mitsumoriSuu" : form.cleaned_data.get('mitsumoriSuu'),
                            "mitsumoriGokei" : mitsumorigokei,
                            "mitsumoriLink" : form.cleaned_data.get('mitsumoriLink'),
                            "ringiTanka" : form.cleaned_data.get('ringiTanka'),
                            "ringiSuu" : form.cleaned_data.get('ringiSuu'),
                            "ringiGokei" : ringigokei,
                            "ringishoNo" : form.cleaned_data.get('ringishoNo'),
                            "ringishoLink" : form.cleaned_data.get('ringishoLink'),
                            "wfNo" : form.cleaned_data.get('wfNo'),
                            "keiyakuKingaku" : form.cleaned_data.get('keiyakuKingaku'),
                            "keiyakushoNo" : form.cleaned_data.get('keiyakushoNo'),
                            "keiyakushoLink" : form.cleaned_data.get('keiyakushoLink'),
                            "onatsuRingiNo" : form.cleaned_data.get('onatsuRingiNo'),
                            "onatsuRingiLink" : form.cleaned_data.get('onatsuRingiLink'),
                            "chumonTanka" : form.cleaned_data.get('chumonTanka'),
                            "chumonSuu" : form.cleaned_data.get('chumonSuu'),
                            "chumonGokei" : chumongokei,
                            "chumonLink" : form.cleaned_data.get('chumonLink'),
                            "nohinKigen" : form.cleaned_data.get('nohinKigen'),
                            "shiharaiKigen" : form.cleaned_data.get('shiharaiKigen'),
                            "nohinTanka" : form.cleaned_data.get('nohinTanka'),
                            "nohinSuu" : form.cleaned_data.get('nohinSuu'),
                            "nohinGokei" : nohingokei,
                            "nohinLink" : form.cleaned_data.get('nohinLink'),
                            "shiharaiTanka" : form.cleaned_data.get('shiharaiTanka'),
                            "konyuSuu" : form.cleaned_data.get('konyuSuu'),
                            "konyuGokei" : konyugokei,
                            "seikyushoLink" : form.cleaned_data.get('seikyushoLink'),
                            "keikaku_Jisseki" : form.cleaned_data.get('keikaku_Jisseki'),
                            "kurikaeshiKigen" : form.cleaned_data.get('kurikaeshiKigen'),
                            "konyuKaisha" : form.cleaned_data.get('konyuKaisha'),
                            # "dataKoshinbi" : form.cleaned_data.get('dataKoshinbi'),
                            "dataKoshinbi" : timezone.now,
                            # "saishuKoshinsha" : form.cleaned_data.get('saishuKoshinsha'),
                            "saishuKoshinsha" : str(request.user),
                            "shainNo" : form.cleaned_data.get('shainNo'),
                            # "tantosha" : form.cleaned_data.get('tantosha'),
                            "tantosha" : tantosha,
                            "comment" : form.cleaned_data.get('comment'),
                            "hyojijun" : hyojijuns
                        }    
                    ) 
            
                
                   
        if checkType == "hukusha":
            if form.is_valid():
                #入力規制
                #空白判定
                if form.cleaned_data.get('keiriShonin') is None:
                    message.append("【経理承認】が空白です")
                if form.cleaned_data.get('statusCode') is None:
                    message.append("【案件ステータスコード】が空白です")
                if form.cleaned_data.get('shiharaiPattern') is None:
                    message.append("【支払パターン】が空白です")
                if form.cleaned_data.get('kanriNo') is None:
                    message.append("【管理番号】が空白です")
                if form.cleaned_data.get('ankenMei') is None:
                    message.append("【案件名】が空白です")
                if form.cleaned_data.get('torihikisakiCode') is None:
                    message.append("【取引先コード】が空白です")
                if form.cleaned_data.get('keikakuTanka') is None:
                    message.append("【計画単価】が空白です")
                if form.cleaned_data.get('keikakuSuu') is None:
                    message.append("【計画数】が空白です")
                #ステータスによる期日空白判定
                if form.cleaned_data.get('statusCode')==9 and form.cleaned_data.get('shiharaiTanka')is None:
                    message.append("【支払単価】が空白です")
                if form.cleaned_data.get('statusCode')==9 and form.cleaned_data.get('konyuSuu')is None:
                    message.append("【購入数】が空白です")

                
                
                if len(message)!=0:
                    pass
            
                else:
                
                    anken = AnkenList.objects.order_by('hyojijun')
                    statuscode = form.cleaned_data.get('statusCode')
                    statusanken = AnkenStatus.objects.filter(ankenStatusCode=int(statuscode))
                    status = statusanken[0].ankenStatus
                    
                    torihikisakicode = form.cleaned_data.get('torihikisakiCode')
                    torihikisakianken = AnkenTorihikisaki.objects.filter(ankenTorihikisakiCode=torihikisakicode)
                    torihikisaki = torihikisakianken[0].ankenTorihikisakiKaisha
                    
                    tantoshacode = form.cleaned_data.get('shainNo')
                    tantoshaanken = AnkenTantosha.objects.filter(ankenTantoshaCode=tantoshacode)
                    tantosha = tantoshaanken[0].ankenTantoshaMei
                    
                    kanjokamokucode = form.cleaned_data.get('kanjokamokuCode')
                    kanjokamokuanken = AnkenKanjokamoku.objects.filter(ankenKanjokamokuCode=kanjokamokucode)
                    kanjokamoku = kanjokamokuanken[0].ankenKanjokamoku

                    keikakugokei=NULL
                    mitsumorigokei=NULL
                    ringigokei  =NULL
                    chumongokei =NULL
                    nohingokei =NULL
                    konyugokei =NULL

                    if form.cleaned_data.get('keikakuTanka')is not None and form.cleaned_data.get('keikakuSuu')is not None:
                        keikakugokei = int(form.cleaned_data.get('keikakuTanka'))*int(form.cleaned_data.get('keikakuSuu'))
                    # if (form.cleaned_data.get('mitsumoriTanka')!=0 or form.cleaned_data.get('mitsumoriTanka') is not None)and (form.cleaned_data.get('mitsumoriSuu')!=0 or form.cleaned_data.get('mitsumoriSuu')is not None):
                    if form.cleaned_data.get('mitsumoriTanka') is not None and form.cleaned_data.get('mitsumoriSuu')is not None:
                        mitsumorigokei = int(form.cleaned_data.get('mitsumoriTanka'))*int(form.cleaned_data.get('mitsumoriSuu'))
                    if form.cleaned_data.get('ringiTanka') is not None and  form.cleaned_data.get('ringiSuu')is not None:                   
                        ringigokei = int(form.cleaned_data.get('ringiTanka'))*int(form.cleaned_data.get('ringiSuu'))
                    if form.cleaned_data.get('chumonTanka')is not None and form.cleaned_data.get('chumonSuu')is not None:
                        chumongokei = int(form.cleaned_data.get('chumonTanka'))*int(form.cleaned_data.get('chumonSuu'))
                    if form.cleaned_data.get('nohinTanka')is not None and form.cleaned_data.get('nohinSuu')is not None:
                        nohingokei = int(form.cleaned_data.get('nohinTanka'))*int(form.cleaned_data.get('nohinSuu'))
                    if form.cleaned_data.get('shiharaiTanka')is not None and form.cleaned_data.get('konyuSuu')is not None:
                        konyugokei = int(form.cleaned_data.get('shiharaiTanka'))*int(form.cleaned_data.get('konyuSuu'))

                    kanriNo=form.cleaned_data.get('kanriNo')
                    edaban=form.cleaned_data.get('edaban')
                    hyojijuns =form.cleaned_data.get('hyojijun')
                    obj =AnkenList.objects.filter(
                        kanriNo= kanriNo,
                        edaban= edaban
                    )    
                    
                    edabanShitei = edaban
                    edabanMax = 0
                    for ankens in anken:
                        if ankens.kanriNo==kanriNo:
                            if ankens.edaban>=edaban:
                                edabanMax=ankens.edaban
                    for ankens in anken:
                        if ankens.kanriNo==kanriNo:
                            if ankens.edaban==edaban:
                                edabanShitei=int(edabanMax)+1
                                
                    # edabanShitei=0
                    for ankens in anken:
                        
                        # if ankens.kanriNo == kanriNo and int(ankens.edaban) > edabanShitei:
                        #     edabanShitei = int(ankens.edaban)
                        if ankens.hyojijun >hyojijuns:
                            ankens.hyojijun=ankens.hyojijun+1
                            ankens.save()
                    
                    for ankens in anken:
                        if ankens.hyojijun == hyojijuns:
                            hyojijuns=hyojijuns+1
                    
                    if int(statuscode) == 99:
                        edabanShitei = 0
                        
                    a = AnkenList(
                            
                            keiriShonin= form.cleaned_data.get('keiriShonin'),
                            statusCode= form.cleaned_data.get('statusCode'),
                            # status=form.cleaned_data.get('status'),
                            status=status,
                            edabanGroup=form.cleaned_data.get('edabanGroup'),
                            shiharaiPattern=form.cleaned_data.get('shiharaiPattern'),
                            kanriNo=form.cleaned_data.get('kanriNo'),
                            edaban=edabanShitei,
                            ankenMei=form.cleaned_data.get('ankenMei'),
                            torihikisakiCode=form.cleaned_data.get('torihikisakiCode'),
                            # torihikisakiMei=form.cleaned_data.get('torihikisakiMei'),
                            torihikisakiMei=torihikisaki,
                            kanjokamokuCode=form.cleaned_data.get('kanjokamokuCode'),
                            kanjokamoku=kanjokamoku,
                            keikakuTanka=form.cleaned_data.get('keikakuTanka'),
                            keikakuSuu=form.cleaned_data.get('keikakuSuu'),
                            keikakuGokei=keikakugokei,
                            mitsumoriTanka=form.cleaned_data.get('mitsumoriTanka'),
                            mitsumoriSuu=form.cleaned_data.get('mitsumoriSuu'),
                            mitsumoriGokei=mitsumorigokei,
                            mitsumoriLink=form.cleaned_data.get('mitsumoriLink'),
                            ringiTanka=form.cleaned_data.get('ringiTanka'),
                            ringiSuu=form.cleaned_data.get('ringiSuu'),
                            ringiGokei=ringigokei,
                            ringishoNo=form.cleaned_data.get('ringishoNo'),
                            ringishoLink=form.cleaned_data.get('ringishoLink'),
                            wfNo=form.cleaned_data.get('wfNo'),
                            keiyakuKingaku=form.cleaned_data.get('keiyakuKingaku'),
                            keiyakushoNo=form.cleaned_data.get('keiyakushoNo'),
                            keiyakushoLink=form.cleaned_data.get('keiyakushoLink'),
                            onatsuRingiNo=form.cleaned_data.get('onatsuRingiNo'),
                            onatsuRingiLink=form.cleaned_data.get('onatsuRingiLink'),
                            chumonTanka=form.cleaned_data.get('chumonTanka'),
                            chumonSuu=form.cleaned_data.get('chumonSuu'),
                            chumonGokei=chumongokei,
                            chumonLink=form.cleaned_data.get('chumonLink'),
                            nohinKigen=form.cleaned_data.get('nohinKigen'),
                            shiharaiKigen=form.cleaned_data.get('shiharaiKigen'),
                            nohinTanka=form.cleaned_data.get('nohinTanka'),
                            nohinSuu=form.cleaned_data.get('nohinSuu'),
                            nohinGokei=nohingokei,
                            nohinLink=form.cleaned_data.get('nohinLink'),
                            shiharaiTanka=form.cleaned_data.get('shiharaiTanka'),
                            konyuSuu=form.cleaned_data.get('konyuSuu'),
                            konyuGokei=konyugokei,
                            seikyushoLink=form.cleaned_data.get('seikyushoLink'),
                            keikaku_Jisseki=form.cleaned_data.get('keikaku_Jisseki'),
                            kurikaeshiKigen=form.cleaned_data.get('kurikaeshiKigen'),
                            konyuKaisha=form.cleaned_data.get('konyuKaisha'),
                            # dataKoshinbi=form.cleaned_data.get('dataKoshinbi'),
                            # dataKoshinbi=timezone.now,
                            dataKoshinbi=datetime.datetime.today(),
                            saishuKoshinsha=str(request.user),
                            shainNo=form.cleaned_data.get('shainNo'),
                            # tantosha=form.cleaned_data.get('tantosha'),
                            tantosha=tantosha,
                            comment=form.cleaned_data.get('comment'),
                            hyojijun=hyojijuns
                    )    
                    
                    a.save()
                
            
                
        if checkType == "sakujo":        
            if form.is_valid():
                anken = AnkenList.objects.order_by('hyojijun')
                obj =AnkenList.objects.filter(
                    kanriNo= form.cleaned_data.get('kanriNo'),
                    edaban= form.cleaned_data.get('edaban')
                )    
                hyojijuns = obj[0].hyojijun
                for ankens in anken:
                    if ankens.hyojijun >hyojijuns:
                        ankens.hyojijun=ankens.hyojijun-1
                        ankens.save()
                
                AnkenList.objects.filter(
                    kanriNo= form.cleaned_data.get('kanriNo'),
                    edaban= form.cleaned_data.get('edaban')
                ).delete() 
    
        i = 1
        anken = AnkenList.objects.order_by('hyojijun')
        for ankens in anken:
            ankens.hyojijun = i
            ankens.save()
            i=i+1
    
    # print("ちぇっくたいぷは",checkType)
    # print("めっせーじは",message)
    return render(
        request,'ankenkanri2/edit3.html',context={'form':form,'ankenList' : AnkenList.objects.order_by('hyojijun'), 'checkType':checkType,'kazu':kazu,'message':message}
    )    